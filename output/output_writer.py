"""
output/output_writer.py

Writes pipeline output buckets to named sheets in an existing .xlsx template.
Each bucket (e.g. "sales_ledger") maps to a sheet of the same name.
Rows are appended after the last occupied row.

Column matching behaviour
--------------------------
If the sheet already has a header row, values are written into the column
whose header matches the DataFrame column name (case-sensitive).  This means
the physical column order in the template is respected and data lands in the
right place regardless of the order columns appear in the DataFrame.

Columns in the DataFrame that have no matching header in the sheet are
appended to the right of the existing headers.
Columns in the template header that have no data are left blank (None).
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook


_DATE_PATTERNS = [
    (re.compile(r"^\d{1,2}/\d{1,2}/\d{4}$"), "%m/%d/%Y"),   # M/D/YYYY
    (re.compile(r"^\d{4}-\d{2}-\d{2}$"),      "%Y-%m-%d"),   # YYYY-MM-DD
]

def _to_dd_mm_yyyy(value: str) -> str:
    for pattern, fmt in _DATE_PATTERNS:
        if pattern.match(value):
            try:
                return datetime.strptime(value, fmt).strftime("%d.%m.%Y")
            except ValueError:
                pass
    return value


_PASSTHROUGH_BUCKETS = {"other", "manual_review"}


def _get_mapped_columns(config, bucket: str) -> list[str]:
    """Return ordered, deduplicated list of output columns for a bucket."""
    tx_config = config.transaction_types.get(bucket)
    if tx_config is None:
        return []

    columns: list[str] = []
    for rule in tx_config.rules:
        rt = rule.rule_type
        if rt == "required_fields":
            columns.extend(rule.params.get("fields", []))
        elif rt in ("direct_mapping", "set_value", "lookup", "compute"):
            col = rule.params.get("target_column")
            if col:
                columns.append(col)
        elif rt == "conditional_mapping":
            for condition in rule.params.get("conditions", []):
                col = condition.get("target_column")
                if col:
                    columns.append(col)

    seen: set[str] = set()
    return [c for c in columns if not (c in seen or seen.add(c))]


def _read_header_map(ws) -> dict[str, int]:
    """
    Scan the first row of *ws* and return {header_name: col_index_1based}.
    Empty / None cells are ignored.
    """
    header_map: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is not None:
            header_map[str(cell.value)] = cell.column
    return header_map


def _resolve_col_map(ws, cols: list[str]) -> dict[str, int]:
    """
    Build a mapping of {df_column_name: excel_col_index} by matching against
    the sheet's existing header row.

    - Matched columns use their existing position in the sheet.
    - Unmatched columns are appended to the right of the last used column.
    """
    header_map = _read_header_map(ws)
    next_col = (max(header_map.values(), default=0) + 1) if header_map else 1

    col_map: dict[str, int] = {}
    for col in cols:
        if col in header_map:
            col_map[col] = header_map[col]
        else:
            # Append new column and write its header into row 1
            ws.cell(row=1, column=next_col, value=col)
            col_map[col] = next_col
            next_col += 1

    return col_map


def write_outputs_to_xlsx(
    outputs: dict[str, pd.DataFrame],
    config,
    template_path: Path,
    output_path: Path,
) -> None:
    """
    Write each non-empty bucket in *outputs* to the matching sheet in the
    xlsx file at *template_path*, then save the result to *output_path*.

    - Sheet is created automatically if missing.
    - Existing headers are matched by name; data lands in the correct column.
    - New columns not present in the template are appended to the right.
    - Temporary columns (prefixed '_') are never written.
    """
    wb: Workbook = load_workbook(template_path)

    for bucket, df in outputs.items():
        if df.empty:
            continue

        sheet_name = bucket
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
            print(f"  [NEW] created sheet '{sheet_name}'")

        # Resolve writable columns
        if bucket in _PASSTHROUGH_BUCKETS:
            cols = [c for c in df.columns if not c.startswith("_")]
        else:
            mapped = _get_mapped_columns(config, bucket)
            cols = [c for c in mapped if c in df.columns and not c.startswith("_")]

        if not cols:
            print(f"  [SKIP] '{bucket}' — no writable columns resolved")
            continue

        ws = wb[sheet_name]
        col_map = _resolve_col_map(ws, cols)   # {col_name: excel_col_index}
        start_row = ws.max_row + 1

        for r_offset, (_, row) in enumerate(df[cols].iterrows()):
            for col_name, value in row.items():
                cell_value = None if pd.isna(value) else value
                if isinstance(cell_value, str):
                    cell_value = _to_dd_mm_yyyy(cell_value)
                ws.cell(
                    row=start_row + r_offset,
                    column=col_map[col_name],
                    value=cell_value,
                )

        print(f"  [{bucket}] {len(df)} row(s) → sheet '{sheet_name}'")

    wb.save(output_path)
    print(f"\nSaved → {output_path}")
