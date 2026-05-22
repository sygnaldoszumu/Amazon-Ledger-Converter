# loading.py
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Low-level helpers — each does exactly one thing
# ---------------------------------------------------------------------------

def extract_row(df: pd.DataFrame, row_index: int) -> list:
    """Return a single DataFrame row as a plain list of values."""
    if row_index < 0 or row_index >= len(df):
        raise IndexError(f"row_index {row_index} is out of range for DataFrame with {len(df)} rows.")
    return df.iloc[row_index].tolist()


def load_or_create_workbook(filepath: str) -> Workbook:
    """Load an existing workbook or create a fresh one if the file is missing."""
    try:
        return load_workbook(filepath)
    except FileNotFoundError:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        return wb


def get_or_create_sheet(wb: Workbook, sheet_name: str) -> Worksheet:
    """Return the named sheet, creating it if it does not exist."""
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    return wb[sheet_name]


def sheet_is_empty(ws: Worksheet) -> bool:
    """Return True when the sheet has no data."""
    return ws.max_row == 0 or (ws.max_row == 1 and ws.cell(1, 1).value is None)


def write_header(ws: Worksheet, columns: list) -> None:
    """Write column names directly into row 1 of the sheet.

    Uses cell-by-cell assignment instead of ws.append() because openpyxl
    reports max_row=1 on a fresh sheet even when row 1 is empty, causing
    append() to land on row 2 and leave a blank row 1.
    """
    for col_idx, name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=name)


def append_row(ws: Worksheet, row: list) -> None:
    """Append a data row to the sheet."""
    ws.append(row)


def save_workbook(wb: Workbook, filepath: str) -> None:
    """Persist the workbook to disk."""
    wb.save(filepath)


# ---------------------------------------------------------------------------
# Orchestrator — composes the helpers, owns no logic of its own
# ---------------------------------------------------------------------------

def save_row_to_xlsx(
    df: pd.DataFrame,
    row_index: int,
    filepath: str,
    sheet_name: str,
    include_header: bool = True,
) -> None:
    """
    Save one row from *df* to *sheet_name* inside *filepath*.

    The file and sheet are created when they do not already exist.
    The row is always appended after existing data; the header is written
    only when the sheet is empty and *include_header* is True.
    """
    row = extract_row(df, row_index)
    wb = load_or_create_workbook(filepath)
    ws = get_or_create_sheet(wb, sheet_name)

    if include_header and sheet_is_empty(ws):
        write_header(ws, list(df.columns))

    append_row(ws, row)
    save_workbook(wb, filepath)

    
